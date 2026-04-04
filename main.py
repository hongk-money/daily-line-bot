import logging
import os
import base64
import json
from datetime import datetime
import pytz
import anthropic
from telegram import Update
from telegram.ext import ApplicationBuilder, MessageHandler, CommandHandler, filters, ContextTypes

BOT_TOKEN = os.environ.get("BOT_TOKEN", "")
CHANNEL_ID = int(os.environ.get("CHANNEL_ID", "0"))
ALLOWED_USER = os.environ.get("ALLOWED_USER", "").lower()
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

DATA_FILE = "line_data.json"

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

def get_hkt_time():
    hkt = pytz.timezone("Asia/Hong_Kong")
    now = datetime.now(hkt)
    return now.strftime("%m.%d %H:%M HKT")

def get_today_hkt():
    hkt = pytz.timezone("Asia/Hong_Kong")
    now = datetime.now(hkt)
    return now.strftime("%Y-%m-%d")

def parse_date_from_caption(caption: str):
    """캡션에서 날짜 파싱 (예: BTC 0324 → 2026-03-24)"""
    import re
    match = re.search(r'\b(\d{4})\b', caption)
    if match:
        mmdd = match.group(1)
        month = int(mmdd[:2])
        day = int(mmdd[2:])
        hkt = pytz.timezone("Asia/Hong_Kong")
        year = datetime.now(hkt).year
        try:
            date_obj = datetime(year, month, day)
            return date_obj.strftime("%Y-%m-%d")
        except:
            return get_today_hkt()
    return get_today_hkt()

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []

def save_data(records):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

def add_record(date: str, coin: str, upper: str, lower: str):
    records = load_data()
    records.append({
        "date": date,
        "coin": coin,
        "upper": upper,
        "lower": lower,
        "recorded_at": get_hkt_time()
    })
    save_data(records)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "✅ 홍크 라인봇 작동 중\n\n"
        "📌 사용법:\n"
        "캡션 형식: [코인] [날짜(선택)]\n\n"
        "예시:\n"
        "ETH → 오늘 날짜 자동\n"
        "BTC 0324 → 3월 24일로 저장\n\n"
        "📊 /excel → 엑셀 파일 다운로드\n"
        "📋 /data → 저장된 데이터 확인"
    )

async def send_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    username = (user.username or "").lower()
    if ALLOWED_USER and username != ALLOWED_USER:
        await update.message.reply_text("❌ 권한 없음")
        return

    records = load_data()
    if not records:
        await update.message.reply_text("📭 저장된 데이터가 없어요.")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "라인 데이터"

        # 헤더 스타일
        header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
        header_font = Font(color="C9A84C", bold=True, size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="333333")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        headers = ["날짜", "코인", "상단 저항선", "하단 지지선", "기록 시간"]
        col_widths = [15, 10, 18, 18, 20]

        for i, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[chr(64+i)].width = w

        ws.row_dimensions[1].height = 28

        # 데이터 정렬 (날짜순)
        sorted_records = sorted(records, key=lambda x: (x.get("date",""), x.get("coin","")))

        btc_fill = PatternFill(start_color="FFF8E7", end_color="FFF8E7", fill_type="solid")
        eth_fill = PatternFill(start_color="E8F4FF", end_color="E8F4FF", fill_type="solid")

        for row_idx, rec in enumerate(sorted_records, 2):
            coin = rec.get("coin", "")
            fill = btc_fill if coin == "BTC" else eth_fill

            values = [
                rec.get("date", ""),
                coin,
                rec.get("upper", ""),
                rec.get("lower", ""),
                rec.get("recorded_at", "")
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
                cell.alignment = center
                cell.border = border

            ws.row_dimensions[row_idx].height = 22

        filename = f"hongk_lines_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(filename)

        with open(filename, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=filename,
                caption=f"📊 총 {len(records)}개 데이터 | {get_hkt_time()}"
            )

        os.remove(filename)

    except Exception as e:
        await update.message.reply_text(f"❌ 엑셀 생성 실패: {str(e)}")
        logger.error(f"엑셀 오류: {e}")

async def show_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    username = (user.username or "").lower()
    if ALLOWED_USER and username != ALLOWED_USER:
        await update.message.reply_text("❌ 권한 없음")
        return

    records = load_data()
    if not records:
        await update.message.reply_text("📭 저장된 데이터가 없어요.")
        return

    sorted_records = sorted(records, key=lambda x: (x.get("date",""), x.get("coin","")))
    text = f"📋 저장된 데이터 ({len(records)}개)\n\n"
    for rec in sorted_records[-20:]:
        coin = rec.get("coin", "")
        emoji = "🟡" if coin == "BTC" else "🔵"
        text += (
            f"{emoji} {rec.get('date','')} {coin}\n"
            f"   상단 {rec.get('upper','')} | 하단 {rec.get('lower','')}\n\n"
        )

    if len(records) > 20:
        text += f"... 외 {len(records)-20}개 (엑셀로 전체 확인)"

    await update.message.reply_text(text)

async def analyze_chart(image_bytes: bytes) -> dict:
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    image_b64 = base64.standard_b64encode(image_bytes).decode("utf-8")

    prompt = """이 차트 이미지에서 두 가지 수평선의 가격값을 읽어주세요.

1. 주황색(오렌지색) 점선 = 상단 저항선
2. 회색 점선 = 하단 지지선

차트 오른쪽 끝에 표시된 숫자값을 읽어주세요.

반드시 아래 형식으로만 답하세요:
상단: [숫자]
하단: [숫자]"""

    message = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=100,
        messages=[{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/jpeg",
                        "data": image_b64,
                    },
                },
                {"type": "text", "text": prompt}
            ],
        }],
    )

    result = message.content[0].text.strip()
    logger.info(f"Claude 분석: {result}")

    upper = lower = None
    for line in result.split('\n'):
        if '상단' in line:
            upper = line.split(':')[-1].strip().replace(',', '')
        elif '하단' in line:
            lower = line.split(':')[-1].strip().replace(',', '')

    return {"upper": upper, "lower": lower}

async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    username = (user.username or "").lower()

    if ALLOWED_USER and username != ALLOWED_USER:
        await update.message.reply_text("❌ 권한 없음")
        return

    caption = (update.message.caption or "").strip()
    caption_upper = caption.upper()

    # 코인 파싱
    if "BTC" in caption_upper or "BITCOIN" in caption_upper:
        coin = "BTC"
        emoji = "🟡"
    elif "ETH" in caption_upper or "ETHEREUM" in caption_upper:
        coin = "ETH"
        emoji = "🔵"
    else:
        await update.message.reply_text("❌ 캡션에 BTC 또는 ETH를 입력해주세요.\n예) BTC 또는 BTC 0324")
        return

    # 날짜 파싱
    date = parse_date_from_caption(caption)
    date_display = date[5:].replace("-", ".")  # 03.24 형식

    processing_msg = await update.message.reply_text(f"⏳ {date_display} {coin} 차트 분석 중...")

    try:
        photo = update.message.photo[-1]
        photo_file = await context.bot.get_file(photo.file_id)
        image_bytes = await photo_file.download_as_bytearray()

        values = await analyze_chart(bytes(image_bytes))

        if values["upper"] and values["lower"]:
            # 데이터 저장
            add_record(date, coin, values["upper"], values["lower"])

            message = (
                f"{emoji} {coin} 주요 라인 | {date_display}\n"
                f"━━━━━━━━━━━━━━━━\n"
                f"🟠 상단 저항선  {values['upper']}\n"
                f"⬜ 하단 지지선  {values['lower']}\n"
                f"━━━━━━━━━━━━━━━━"
            )

            await context.bot.send_message(chat_id=CHANNEL_ID, text=message)
            await processing_msg.edit_text(f"✅ 채널 발송 + 데이터 저장 완료!\n\n{message}")
        else:
            await processing_msg.edit_text("⚠️ 라인값 추출 실패. 다시 시도해주세요.")

    except Exception as e:
        await processing_msg.edit_text(f"❌ 오류: {str(e)}")
        logger.error(f"오류: {e}")

def main():
    logger.info(f"봇 시작 - CHANNEL_ID: {CHANNEL_ID}, ALLOWED_USER: {ALLOWED_USER}")
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("excel", send_excel))
    app.add_handler(CommandHandler("data", show_data))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
